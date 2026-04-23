"""
NLP_test1.py
============
Financial‑news NLP sentiment classifier for SP500 headlines.

Reads  : ../datasets/SP500V3_editable_test.xlsx  (column "news_headlines")
Writes : same file, new column "NLP_predict" with values: buy / sell / hold

Approach
--------
1.  A curated, finance‑domain lexicon of ~300 positive and negative terms
    weighted by market‑relevance (e.g. "surge" is strongly positive,
    "recall" is mildly negative).
2.  Each headline is tokenised (lowercased, punctuation stripped).
3.  Negation handling: a simple negation window flips the polarity of the
    next 3 tokens after a negation cue ("not", "no", "never", …).
4.  Per‑headline sentiment score = sum of matched‑token weights.
5.  Aggregate score for a row = mean of per‑headline scores.
6.  Thresholding:  score >  +T  → "buy"
                   score < −T  → "sell"
                   otherwise   → "hold"
    where T is calibrated on the dataset's score distribution.

No external model downloads required — runs offline.
"""

import os, re, string, statistics
import openpyxl

# ──────────────────────────────────────────────────────────────────────
# 1.  FINANCIAL SENTIMENT LEXICON
# ──────────────────────────────────────────────────────────────────────

# Positive terms  (word → weight)
_POS = {
    # Strong positive (+2)
    "surge": 2, "surges": 2, "surging": 2, "soar": 2, "soars": 2,
    "soaring": 2, "boom": 2, "booming": 2, "skyrocket": 2, "rally": 2,
    "rallies": 2, "rallying": 2, "breakthrough": 2, "record-high": 2,
    "outperform": 2, "outperforms": 2, "bullish": 2, "blockbuster": 2,
    "blowout": 2, "robust": 2,

    # Moderate positive (+1)
    "gain": 1, "gains": 1, "gaining": 1, "rise": 1, "rises": 1,
    "rising": 1, "rose": 1, "climb": 1, "climbs": 1, "climbing": 1,
    "up": 1, "higher": 1, "high": 1, "growth": 1, "growing": 1,
    "grow": 1, "grew": 1, "expand": 1, "expands": 1, "expanding": 1,
    "expansion": 1, "profit": 1, "profits": 1, "profitable": 1,
    "profitability": 1, "revenue": 1, "revenues": 1, "earnings": 1,
    "beat": 1, "beats": 1, "beating": 1, "exceed": 1, "exceeds": 1,
    "exceeded": 1, "exceeding": 1, "strong": 1, "strength": 1,
    "strengthen": 1, "strengthens": 1, "upgrade": 1, "upgrades": 1,
    "upgraded": 1, "optimism": 1, "optimistic": 1, "upbeat": 1,
    "positive": 1, "improve": 1, "improves": 1, "improved": 1,
    "improvement": 1, "recover": 1, "recovers": 1, "recovery": 1,
    "recovering": 1, "rebound": 1, "rebounds": 1, "boost": 1,
    "boosts": 1, "boosted": 1, "advance": 1, "advances": 1,
    "advancing": 1, "upturn": 1, "upside": 1, "innovation": 1,
    "innovative": 1, "launch": 1, "launches": 1, "launched": 1,
    "opportunity": 1, "opportunities": 1, "dividend": 1, "dividends": 1,
    "buyback": 1, "buybacks": 1, "acquisition": 1, "merger": 1,
    "approval": 1, "approved": 1, "approves": 1, "success": 1,
    "successful": 1, "win": 1, "wins": 1, "winner": 1, "winning": 1,
    "record": 1, "milestone": 1, "confidence": 1, "confident": 1,
    "momentum": 1, "outpace": 1, "demand": 1, "hire": 1, "hiring": 1,
    "hires": 1, "jobs": 1, "employment": 1, "resilient": 1,
    "resilience": 1, "impressive": 1, "impresses": 1, "stellar": 1,
    "value": 1, "upward": 1, "benefit": 1, "benefits": 1,
    "favourable": 1, "favorable": 1, "supercomputer": 1,
    "ai": 1, "artificial intelligence": 1,
}

# Negative terms  (word → weight, stored as positive, applied as negative)
_NEG = {
    # Strong negative (−2)
    "crash": 2, "crashes": 2, "crashing": 2, "plunge": 2, "plunges": 2,
    "plunging": 2, "plummet": 2, "plummets": 2, "plummeting": 2,
    "collapse": 2, "collapses": 2, "collapsing": 2, "crisis": 2,
    "recession": 2, "bankruptcy": 2, "bankrupt": 2, "default": 2,
    "defaults": 2, "bearish": 2, "meltdown": 2, "panic": 2,
    "catastrophe": 2, "catastrophic": 2, "devastate": 2,
    "devastating": 2, "turmoil": 2, "freefall": 2,

    # Moderate negative (−1)
    "fall": 1, "falls": 1, "falling": 1, "fell": 1, "drop": 1,
    "drops": 1, "dropping": 1, "dropped": 1, "decline": 1,
    "declines": 1, "declining": 1, "decrease": 1, "decreases": 1,
    "low": 1, "lower": 1, "lowest": 1, "down": 1, "downturn": 1,
    "downgrade": 1, "downgrades": 1, "downgraded": 1, "loss": 1,
    "losses": 1, "lose": 1, "loses": 1, "losing": 1, "lost": 1,
    "miss": 1, "misses": 1, "missed": 1, "missing": 1, "weak": 1,
    "weaken": 1, "weakens": 1, "weakness": 1, "negative": 1,
    "pessimism": 1, "pessimistic": 1, "concern": 1, "concerns": 1,
    "worried": 1, "worry": 1, "worries": 1, "fear": 1, "fears": 1,
    "risk": 1, "risks": 1, "risky": 1, "threat": 1, "threatens": 1,
    "threatening": 1, "uncertainty": 1, "uncertain": 1, "volatile": 1,
    "volatility": 1, "slump": 1, "slumps": 1, "stall": 1, "stalls": 1,
    "stalling": 1, "selloff": 1, "sell-off": 1, "selling": 1,
    "layoff": 1, "layoffs": 1, "cut": 1, "cuts": 1, "cutting": 1,
    "slash": 1, "slashes": 1, "debt": 1, "deficit": 1, "inflation": 1,
    "tariff": 1, "tariffs": 1, "sanction": 1, "sanctions": 1,
    "fraud": 1, "scandal": 1, "lawsuit": 1, "sue": 1, "sued": 1,
    "fine": 1, "fined": 1, "penalty": 1, "violation": 1, "shutdown": 1,
    "slowdown": 1, "slowing": 1, "slow": 1, "stagnant": 1,
    "stagnation": 1, "contraction": 1, "shrink": 1, "shrinks": 1,
    "war": 1, "conflict": 1, "attack": 1, "terror": 1, "terrorism": 1,
    "gloom": 1, "gloomy": 1, "bust": 1, "trouble": 1, "troubled": 1,
    "struggles": 1, "struggling": 1, "underperform": 1,
    "underperforms": 1, "overvalued": 1, "bubble": 1,
    "manipulation": 1, "glut": 1, "oversupply": 1, "dumping": 1,
    "exodus": 1, "flee": 1, "outage": 1, "recall": 1, "recalls": 1,
    "oust": 1, "ousted": 1, "resign": 1, "resigns": 1, "resigned": 1,
    "impeach": 1, "ban": 1, "bans": 1, "banned": 1,
}

# Negation cues
_NEGATION_CUES = frozenset([
    "not", "no", "never", "neither", "nobody", "nothing", "nowhere",
    "nor", "cannot", "can't", "couldn't", "didn't", "doesn't", "don't",
    "hadn't", "hasn't", "haven't", "isn't", "wasn't", "weren't",
    "won't", "wouldn't", "hardly", "barely", "scarcely",
])

_NEGATION_WINDOW = 3          # flip polarity for N tokens after negation cue
_HEADLINE_SEP = re.compile(r"\s*;\s*")
_TOKEN_RE = re.compile(r"[a-z][a-z'\-]+")


# ──────────────────────────────────────────────────────────────────────
# 2.  SCORING FUNCTIONS
# ──────────────────────────────────────────────────────────────────────

def _score_headline(headline: str) -> float:
    """Return a sentiment score for a single headline string."""
    tokens = _TOKEN_RE.findall(headline.lower())
    score = 0.0
    neg_remaining = 0          # countdown for negation window

    for tok in tokens:
        if tok in _NEGATION_CUES:
            neg_remaining = _NEGATION_WINDOW
            continue

        flip = -1 if neg_remaining > 0 else 1

        if tok in _POS:
            score += flip * _POS[tok]
        elif tok in _NEG:
            score -= flip * _NEG[tok]

        if neg_remaining > 0:
            neg_remaining -= 1

    return score


def score_row(news_cell: str) -> float:
    """
    Score an entire cell of semicolon‑separated headlines.
    Returns the *mean* per‑headline score (normalises for rows with
    different numbers of headlines).
    """
    if not news_cell or not isinstance(news_cell, str) or news_cell.strip() == "":
        return 0.0
    headlines = [h.strip() for h in _HEADLINE_SEP.split(news_cell) if h.strip()]
    if not headlines:
        return 0.0
    scores = [_score_headline(h) for h in headlines]
    return statistics.mean(scores)


# ──────────────────────────────────────────────────────────────────────
# 3.  CLASSIFY ROWS  →  buy / sell / hold
# ──────────────────────────────────────────────────────────────────────

def classify(score: float, buy_threshold: float, sell_threshold: float) -> str:
    if score > buy_threshold:
        return "buy"
    else:
        return "sell"

# ──────────────────────────────────────────────────────────────────────
# 4.  MAIN  –  read → score → threshold → write
# ──────────────────────────────────────────────────────────────────────

def main():
    # ── Paths ────────────────────────────────────────────────────────
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(script_dir)
    dataset_path = os.path.join(project_dir, "datasets",
                                "SP500V3_editable_test.xlsx")

    print(f"[NLP_test1] Loading workbook: {dataset_path}")
    wb = openpyxl.load_workbook(dataset_path)
    ws = wb.active

    # ── Locate columns ──────────────────────────────────────────────
    header = [cell.value for cell in ws[1]]
    try:
        news_col_idx = header.index("news_headlines") + 1   # 1‑based
    except ValueError:
        raise SystemExit("ERROR: column 'news_headlines' not found in the sheet.")

    # Determine the column for NLP_predict (append after last col)
    predict_col_idx = ws.max_column + 1
    ws.cell(row=1, column=predict_col_idx, value="NLP_predict")

    # ── Score every row ──────────────────────────────────────────────
    total_rows = ws.max_row
    scores = []

    print(f"[NLP_test1] Scoring {total_rows - 1} data rows …")
    for row_num in range(2, total_rows + 1):
        cell_val = ws.cell(row=row_num, column=news_col_idx).value
        s = score_row(cell_val)
        scores.append((row_num, s))

    # ── Adaptive thresholding ────────────────────────────────────────
    # Use percentiles of the score distribution so the buy/sell/hold
    # ratio adapts to the dataset rather than being hard‑coded.
    all_scores = [s for _, s in scores]
    sorted_scores = sorted(all_scores)
    n = len(sorted_scores)

    # 33rd and 67th percentiles → roughly equal class splits
    p33 = sorted_scores[int(n * 0.33)]
    p67 = sorted_scores[int(n * 0.67)]

    # Ensure thresholds are symmetric around 0 when distribution allows
    buy_thresh = max(p67, 0.05)
    sell_thresh = min(p33, -0.05)

    print(f"[NLP_test1] Thresholds  →  buy > {buy_thresh:.3f}  |  "
          f"sell < {sell_thresh:.3f}")

    # ── Write predictions ────────────────────────────────────────────
    buy_count = sell_count = hold_count = 0
    for row_num, s in scores:
        label = classify(s, buy_thresh, sell_thresh)
        ws.cell(row=row_num, column=predict_col_idx, value=label)
        if label == "buy":
            buy_count += 1
        else:
            sell_count += 1
        # else:
        #     hold_count += 1

    # ── Save workbook ────────────────────────────────────────────────
    print(f"[NLP_test1] Saving workbook …")
    wb.save(dataset_path)
    print(f"[NLP_test1] Done.  Predictions written to column "
          f"'{ws.cell(row=1, column=predict_col_idx).value}'")
    # print(f"            buy={buy_count}  sell={sell_count}  hold={hold_count}  "
    #       f"total={buy_count + sell_count + hold_count}")


if __name__ == "__main__":
    main()
